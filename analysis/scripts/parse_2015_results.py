"""
Parse 2015 Alberta provincial election results into CSV.

Source: data/2015_results.xlsx (Elections Alberta 2015PGE-Official-Results.xlsx).

Output: data/alberta_2015_results.csv with per-ED totals for NDP, PC,
WRP, LIB, and other parties. Also computes the combined PC+WRP total
(the 2017 merger that became UCP), for cross-election comparison.

Boundary caveat: 2015 EDs used the 2010 commission's boundaries, which
differ from 2019's. Name matches between 2015 and 2019 are direct where
possible; boundary changes are noted in the output.
"""
from __future__ import annotations
import csv
import os
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent.parent
DATA = ROOT / "data"


# Classify parties into cross-election groups.
# UCP (2019+) = PC + WRP merger (2017).
# "Other" lumps every small party (GPA, AP, SC, CP-A, LIB, IND, etc.).
PARTY_NDP = {"NDP"}
PARTY_PC = {"PC"}
PARTY_WRP = {"WRP"}
PARTY_LIB = {"LIB"}


def extract_party_from_header(cell: str) -> str | None:
    """Cell format is 'FIRSTNAME\\nLASTNAME\\nPARTY'. Return PARTY."""
    if not cell:
        return None
    parts = [p.strip() for p in str(cell).split("\n") if p.strip()]
    if not parts:
        return None
    return parts[-1].upper()


def clean_ed_name(raw: str) -> str:
    """Strip 'Summary Of Results By Poll - ' prefix and trailing spaces."""
    if not raw:
        return ""
    m = re.match(r"Summary Of Results By Poll - (.+)$", raw.strip(), re.IGNORECASE)
    return (m.group(1) if m else raw).strip()


def parse_sheet(sh) -> dict:
    """Parse one 2015 ED sheet into a dict of party totals."""
    title = clean_ed_name(sh["A1"].value or "")
    header_row = [c.value for c in sh[2]]

    # Locate candidate columns — those whose header is 'FIRST\nLAST\nPARTY'.
    # Must have 3+ lines (First, Last, Party). 'Names\non List' has 2 lines so
    # it's filtered out. Short headers like 'V', 'S', 'D', 'R' are single-line.
    cand_cols = {}  # col_idx -> party
    for idx, cell in enumerate(header_row):
        if not isinstance(cell, str):
            continue
        parts = [p.strip() for p in cell.split("\n") if p.strip()]
        if len(parts) < 3:
            continue
        party = parts[-1].upper()
        # Valid party codes are 2-5 letters; filter anything else
        if not (2 <= len(party) <= 5 and party.isalpha() or "-" in party):
            continue
        cand_cols[idx] = party

    # Sum votes per party. Each data row is a poll.
    totals = {}
    rows_parsed = 0
    for row in sh.iter_rows(min_row=3, values_only=True):
        if not row or row[0] is None:
            continue
        # Skip summary rows at the bottom (usually text in col A)
        if isinstance(row[0], str) and not row[0].strip().isdigit() and row[0] != "-":
            continue
        rows_parsed += 1
        for idx, party in cand_cols.items():
            if idx < len(row) and isinstance(row[idx], (int, float)):
                totals[party] = totals.get(party, 0) + int(row[idx])

    return {
        "ed_2015": title,
        "rows_parsed": rows_parsed,
        "party_totals": totals,
    }


def main():
    wb = openpyxl.load_workbook(DATA / "2015_results.xlsx", data_only=True)
    results = []
    for sheet_name in wb.sheetnames:
        sh = wb[sheet_name]
        parsed = parse_sheet(sh)
        if not parsed["ed_2015"]:
            continue
        totals = parsed["party_totals"]
        ndp = totals.get("NDP", 0)
        pc = totals.get("PC", 0)
        wrp = totals.get("WRP", 0)
        lib = totals.get("LIB", 0)
        ucp_equiv = pc + wrp  # 2017 merger to UCP
        other_cols = {k: v for k, v in totals.items()
                      if k not in {"NDP", "PC", "WRP", "LIB"}}
        other_total = sum(other_cols.values())
        total = ndp + pc + wrp + lib + other_total
        results.append({
            "sheet": sheet_name,
            "ed_2015": parsed["ed_2015"],
            "ndp": ndp,
            "pc": pc,
            "wrp": wrp,
            "ucp_equiv": ucp_equiv,
            "lib": lib,
            "other": other_total,
            "total": total,
            "rows_parsed": parsed["rows_parsed"],
        })

    # Sanity-check: province-wide totals
    tot_ndp = sum(r["ndp"] for r in results)
    tot_pc = sum(r["pc"] for r in results)
    tot_wrp = sum(r["wrp"] for r in results)
    tot_lib = sum(r["lib"] for r in results)
    tot_ucp_eq = sum(r["ucp_equiv"] for r in results)
    grand = sum(r["total"] for r in results)
    print(f"Parsed {len(results)} EDs, {sum(r['rows_parsed'] for r in results):,} poll rows")
    print(f"Totals: NDP {tot_ndp:,}  PC {tot_pc:,}  WRP {tot_wrp:,}  LIB {tot_lib:,}")
    print(f"UCP-equivalent (PC+WRP): {tot_ucp_eq:,}")
    print(f"Grand total: {grand:,}")
    print(f"NDP two-party vs UCP-equiv: {tot_ndp/(tot_ndp+tot_ucp_eq)*100:.2f}%")

    # Reference: official 2015 popular vote was NDP 40.59% / WRP 24.22% /
    # PC 27.79% / LIB 4.18% of total valid. Our parse should roughly match.
    print("\nFor reference (2015 official):")
    print("  NDP 40.59%, PC 27.79%, WRP 24.22%, LIB 4.18%, other 3.22%")
    print(f"Parsed shares:")
    if grand:
        print(f"  NDP {tot_ndp/grand*100:.2f}%, PC {tot_pc/grand*100:.2f}%, "
              f"WRP {tot_wrp/grand*100:.2f}%, LIB {tot_lib/grand*100:.2f}%")

    # Write CSV
    out_path = DATA / "alberta_2015_results.csv"
    with open(out_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["sheet", "ed_2015", "ndp", "pc", "wrp", "ucp_equiv",
                    "lib", "other", "total", "rows_parsed"])
        for r in results:
            w.writerow([r["sheet"], r["ed_2015"], r["ndp"], r["pc"], r["wrp"],
                        r["ucp_equiv"], r["lib"], r["other"],
                        r["total"], r["rows_parsed"]])
    print(f"\nWrote {out_path}")


if __name__ == "__main__":
    main()
