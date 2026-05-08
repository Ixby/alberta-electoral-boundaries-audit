"""
Alberta Electoral Boundaries Audit — Poll-Location Attribution Skeleton
========================================================================
v0.1 — Starting point for Phase 4C of the audit.

This script does the data loading, parsing, and structuring for the
poll-location-based vote attribution method. It does NOT do the geocoding,
spatial assignment, or vote apportionment — those are the next session's
work, with this skeleton as the starting framework.

USAGE:
    python3 electoral_forensics_poll_attribution_skeleton.py

This will print a summary of the parsed Statement of Vote data and
report what's loaded vs what still needs implementation.

DEPENDENCIES (from setup.sh):
    - pandas, openpyxl  (parsing)
    - geopy             (Nominatim fallback geocoding)
    - requests          (downloading Statement of Vote)

PIPELINE STAGES (this skeleton implements 1-2; 3-7 are stubbed):
    1. Download Statement of Vote xlsx if not present
    2. Parse all 87 sheets into a unified poll-level dataframe
    3. Build landmark dictionary (school district lists, etc.)  [STUB]
    4. Geocode polls via dictionary + Nominatim fallback        [STUB]
    5. Zero-Sum Verification — check geocoded polls fall in 2019 ED [STUB]
    6. Vision-based assignment to 2026 hybrid EDs               [STUB]
    7. Apportion Advance/Mobile/Special by Election Day share   [STUB]

CRITICAL METHODOLOGY (read before extending):
    Vote Anywhere (2023 election): 21.9% of voters used advance/special
    ballots that could be cast at any location. Votes are attributed to
    voter HOME ED, not physical location. This means:

    - Election Day polls = spatially valid (voter lives near here)
    - Advance/Mobile/Special polls = NOT spatially valid (voter could be
      anywhere); already home-ED-attributed

    Use Election Day for spatial assignment. Apportion the rest by the
    Election Day spatial share within each contributing 2019 ED.
"""

# Version: 0.1 series  (last updated 2026-04-26)


import os
import sys
import csv
import re
from pathlib import Path
from typing import Dict, List, Optional

# ED order in the Statement of Vote sheets (sheet "01" through "87")
# Calgary 1-26 alphabetical (Bhullar-McCall sorts as McCall historically)
# Edmonton 27-46 alphabetical
# Rest of Alberta 47-87 alphabetical
ED_NAMES_2019 = [
    # Calgary 1-26
    "Calgary-Acadia",
    "Calgary-Beddington",
    "Calgary-Bow",
    "Calgary-Buffalo",
    "Calgary-Cross",
    "Calgary-Currie",
    "Calgary-East",
    "Calgary-Edgemont",
    "Calgary-Elbow",
    "Calgary-Falconridge",
    "Calgary-Fish Creek",
    "Calgary-Foothills",
    "Calgary-Glenmore",
    "Calgary-Hays",
    "Calgary-Klein",
    "Calgary-Lougheed",
    "Calgary-Bhullar-McCall",
    "Calgary-Mountain View",
    "Calgary-North",
    "Calgary-North East",
    "Calgary-North West",
    "Calgary-Peigan",
    "Calgary-Shaw",
    "Calgary-South East",
    "Calgary-Varsity",
    "Calgary-West",
    # Edmonton 27-46
    "Edmonton-Beverly-Clareview",
    "Edmonton-Castle Downs",
    "Edmonton-City Centre",
    "Edmonton-Decore",
    "Edmonton-Ellerslie",
    "Edmonton-Glenora",
    "Edmonton-Gold Bar",
    "Edmonton-Highlands-Norwood",
    "Edmonton-Manning",
    "Edmonton-McClung",
    "Edmonton-Meadows",
    "Edmonton-Mill Woods",
    "Edmonton-North West",
    "Edmonton-Riverview",
    "Edmonton-Rutherford",
    "Edmonton-South",
    "Edmonton-South West",
    "Edmonton-Strathcona",
    "Edmonton-West Henday",
    "Edmonton-Whitemud",
    # Rest of Alberta 47-87
    "Airdrie-Cochrane",
    "Airdrie-East",
    "Athabasca-Barrhead-Westlock",
    "Banff-Kananaskis",
    "Bonnyville-Cold Lake-St. Paul",
    "Brooks-Medicine Hat",
    "Camrose",
    "Cardston-Siksika",
    "Central Peace-Notley",
    "Chestermere-Strathmore",
    "Cypress-Medicine Hat",
    "Drayton Valley-Devon",
    "Drumheller-Stettler",
    "Fort McMurray-Lac La Biche",
    "Fort McMurray-Wood Buffalo",
    "Fort Saskatchewan-Vegreville",
    "Grande Prairie",
    "Grande Prairie-Wapiti",
    "Highwood",
    "Innisfail-Sylvan Lake",
    "Lac Ste. Anne-Parkland",
    "Lacombe-Ponoka",
    "Leduc-Beaumont",
    "Lesser Slave Lake",
    "Lethbridge-East",
    "Lethbridge-West",
    "Livingstone-Macleod",
    "Maskwacis-Wetaskiwin",
    "Morinville-St. Albert",
    "Olds-Didsbury-Three Hills",
    "Peace River",
    "Red Deer-North",
    "Red Deer-South",
    "Rimbey-Rocky Mountain House-Sundre",
    "Sherwood Park",
    "Spruce Grove-Stony Plain",
    "St. Albert",
    "Strathcona-Sherwood Park",
    "Taber-Warner",
    "Vermilion-Lloydminster-Wainwright",
    "West Yellowhead",
]
assert len(ED_NAMES_2019) == 87


def find_or_download_statement_of_vote() -> str:
    """Locate the Statement of Vote xlsx, or download it if not present."""
    here = Path(__file__).parent
    candidates = [
        here / "2023_results.xlsx",
        here.parent / "data" / "2023_results.xlsx",
        here.parent / "2023_results.xlsx",
    ]
    for c in candidates:
        if c.exists():
            print(f"Found Statement of Vote: {c}")
            return str(c)

    # Download
    url = "https://www.elections.ab.ca/uploads/2023-Provincial-General-Election-Statement-of-Vote.xlsx"
    target = here / "2023_results.xlsx"
    print(f"Downloading Statement of Vote from {url}")
    print(f"  Target: {target}")
    try:
        import urllib.request

        urllib.request.urlretrieve(url, target)
        size = target.stat().st_size
        print(f"  Downloaded: {size:,} bytes")
        return str(target)
    except Exception as e:
        print(f"  Download failed: {e}")
        print(f"  Manual fallback: download from {url} and place at {target}")
        sys.exit(1)


def parse_statement_of_vote(xlsx_path: str) -> List[Dict]:
    """Parse all 87 sheets into a unified poll-level dataframe (as list of dicts).

    Returns: list of poll records with keys:
        ed_2019, sheet_num, poll_letter, poll_name, ballot_type,
        ndp_votes, ucp_votes, other_votes, valid_votes, voting_areas
    """
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    polls = []

    for sheet_idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        ed_name = ED_NAMES_2019[sheet_idx]

        # Find candidate columns from header row 1
        header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        ndp_col = ucp_col = valid_col = type_col = name_col = areas_col = None
        other_cols = []

        for ci, h in enumerate(header):
            if not h or not isinstance(h, str):
                continue
            if h == "Voting Location":
                # Note: the first col is poll letter (A, B, C); col 2 is poll name
                pass
            elif h == "Type":
                type_col = ci
            elif h == "Voting Areas Assigned":
                areas_col = ci
            elif h == "Valid Ballots Cast":
                valid_col = ci
            elif "\n" in h:
                # Candidate column: "Name\nPARTY"
                if h.endswith("\nNDP"):
                    ndp_col = ci
                elif h.endswith("\nUCP"):
                    ucp_col = ci
                else:
                    other_cols.append(ci)

        # Iterate data rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                # Subtotal/blank row — check by looking at col 2
                if row[1] and isinstance(row[1], str):
                    if (
                        "Subtotal" in row[1]
                        or "Total" in row[1]
                        or "Percentage" in row[1]
                    ):
                        continue  # skip aggregation rows
                else:
                    continue

            # Extract poll-level data
            poll_letter = row[0] if row[0] and isinstance(row[0], str) else None
            poll_name = row[1] if len(row) > 1 else None
            ballot_type = (
                row[type_col] if type_col is not None and type_col < len(row) else None
            )
            voting_areas = (
                row[areas_col]
                if areas_col is not None and areas_col < len(row)
                else None
            )
            ndp_votes = (
                row[ndp_col] if ndp_col is not None and ndp_col < len(row) else 0
            )
            ucp_votes = (
                row[ucp_col] if ucp_col is not None and ucp_col < len(row) else 0
            )
            valid_votes = (
                row[valid_col] if valid_col is not None and valid_col < len(row) else 0
            )
            other_votes = sum(row[c] or 0 for c in other_cols if c < len(row))

            if not poll_name or not isinstance(poll_name, str):
                continue

            polls.append(
                {
                    "ed_2019": ed_name,
                    "sheet_num": sheet_name,
                    "poll_letter": poll_letter,
                    "poll_name": (
                        poll_name.strip() if isinstance(poll_name, str) else poll_name
                    ),
                    "ballot_type": ballot_type,
                    "ndp_votes": int(ndp_votes or 0),
                    "ucp_votes": int(ucp_votes or 0),
                    "other_votes": int(other_votes or 0),
                    "valid_votes": int(valid_votes or 0),
                    "voting_areas": voting_areas,
                    "lat": None,  # filled in geocoding stage
                    "lon": None,
                    "ed_2026_majority": None,  # filled in spatial assignment
                    "ed_2026_minority": None,
                }
            )

    return polls


def geocode_via_landmark_dictionary(polls: List[Dict]) -> List[Dict]:
    """STUB: Match poll names against an Alberta landmarks dictionary.

    Implementation strategy:
    1. Load `analysis/alberta_landmarks.csv` if it exists, columns:
       landmark_type, name, latitude, longitude, city, address
    2. Sources to build the dictionary from:
       - Calgary Board of Education school list with addresses
       - Edmonton Public Schools list
       - Calgary/Edmonton Catholic School District lists
       - Alberta Education provincial school directory
       - City of Calgary community association directory
       - City of Edmonton community recreation centres list
    3. Match poll_name fuzzy against landmark name (use rapidfuzz or
       difflib.SequenceMatcher with cutoff ~85)
    4. On match: copy lat/lon to poll record. Mark match_method='landmark_dict'

    Returns the polls list with lat/lon filled where matched, None otherwise.
    """
    print("[STUB] geocode_via_landmark_dictionary — implement before running")
    return polls


def geocode_via_nominatim(polls: List[Dict]) -> List[Dict]:
    """STUB: Geocode unmatched polls via Nominatim with 1-call/sec rate limit.

    Implementation strategy:
    1. Filter polls where lat is None (didn't match landmark dict)
    2. Build query: f"{poll_name}, {city}, AB, Canada" — infer city from ed_2019
    3. Call geopy.geocoders.Nominatim(user_agent='alberta_audit_v0.7')
    4. Sleep 1 second between calls (Nominatim policy)
    5. Cache results to analysis/poll_geocoding_cache.csv
    6. On retry: load cache first, only call API for new polls

    Returns the polls list with lat/lon filled where geocoded.
    """
    print("[STUB] geocode_via_nominatim — implement before running")
    return polls


def zero_sum_verification(polls: List[Dict], boundaries_2019_geojson: str) -> Dict:
    """STUB: Verify each geocoded poll falls within its known 2019 ED.

    Implementation strategy:
    1. Load 2019 ED boundary geometry (from Elections Alberta or OSM-derived)
    2. For each poll with lat/lon:
       - Build shapely.geometry.Point(lon, lat)
       - Check Point.within(polygon for ed_2019)
    3. Flag mismatches. Return summary:
       {'total': N, 'verified': V, 'failed': F, 'failed_polls': [list]}
    4. Failed polls: re-geocode manually, discard, or accept depending on rate
    5. Hard gate: if failure rate > 10%, the geocoding pipeline isn't trustworthy
    """
    print("[STUB] zero_sum_verification — implement before running")
    return {}


def assign_to_2026_eds_via_vision(
    polls: List[Dict], map_image_paths: Dict[str, str]
) -> List[Dict]:
    """STUB: For each Election Day poll, determine which new 2026 ED contains it.

    Implementation strategy:
    1. Filter polls where ballot_type == 'Election Day' and lat/lon present
    2. For each poll in a hybrid-affected region (Calgary, Edmonton, hybrid-near-cities):
       - Load the appropriate map JPG
       - Use Claude vision to determine which numbered ED contains the (lat, lon)
       - Both for majority and minority maps
    3. For polls in non-hybrid regions: map 1:1 from 2019 to 2026 (same ED, possibly renamed)
    4. Document borderline cases in analysis/borderline_poll_resolution.md

    Returns polls list with ed_2026_majority and ed_2026_minority filled.
    """
    print("[STUB] assign_to_2026_eds_via_vision — implement before running")
    return polls


def apportion_special_ballots(polls: List[Dict]) -> List[Dict]:
    """STUB: Distribute Advance/Mobile/Special votes by Election Day spatial share.

    Implementation strategy:
    1. Group Election Day polls by (ed_2019, ed_2026_minority) — same for majority
    2. For each 2019 ED: compute fraction of Election Day votes that went to each new 2026 ED
       e.g., Calgary-Bow had 100% Election Day votes go to Calgary-Bow-Springbank → 1.00
       e.g., Calgary-Edgemont had 55% go to Nolan Hill-Cochrane, 45% stay in Edgemont
    3. For each Advance/Mobile/Special row in the original 2019 ED:
       - Distribute votes to new 2026 EDs by the computed fraction
    4. Add these apportioned votes to the new 2026 ED totals

    Returns polls list with apportioned special ballot rows added.
    """
    print("[STUB] apportion_special_ballots — implement before running")
    return polls


def main():
    print("Alberta Boundaries Audit — Poll-Location Attribution (v0.1 skeleton)")
    print("=" * 70)

    xlsx_path = find_or_download_statement_of_vote()
    print(f"\nParsing Statement of Vote...")
    polls = parse_statement_of_vote(xlsx_path)

    # Summary stats
    n_total = len(polls)
    by_type = {}
    for p in polls:
        bt = p["ballot_type"] or "Unknown"
        by_type[bt] = by_type.get(bt, 0) + 1

    print(f"\nParsed {n_total} poll records across 87 EDs")
    print(f"  By ballot type:")
    for bt, n in sorted(by_type.items(), key=lambda x: -x[1]):
        print(f"    {bt:20s}: {n}")

    # Verify totals against known result
    total_ndp = sum(p["ndp_votes"] for p in polls)
    total_ucp = sum(p["ucp_votes"] for p in polls)
    print(f"\n  Total NDP votes parsed: {total_ndp:,}")
    print(f"  Total UCP votes parsed: {total_ucp:,}")
    print(f"  Two-party total:         {total_ndp + total_ucp:,}")
    print(f"  Expected (per Wikipedia, two-party-ish): ~1.7M")

    # Save the parsed dataframe for the next pipeline stages
    out_csv = Path(__file__).parent / "polls_2023_unified.csv"
    with open(out_csv, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(polls[0].keys()))
        writer.writeheader()
        writer.writerows(polls)
    print(f"\nSaved unified poll dataframe: {out_csv}")
    print(f"  Rows: {n_total}")
    print(f"  Columns: {', '.join(polls[0].keys())}")

    # Pipeline stage stubs — implement these
    print(f"\n{'─' * 70}")
    print("Pipeline stages requiring implementation:")
    polls = geocode_via_landmark_dictionary(polls)
    polls = geocode_via_nominatim(polls)
    verification_report = zero_sum_verification(polls, "boundaries_2019.geojson")
    polls = assign_to_2026_eds_via_vision(
        polls,
        {
            "majority_calgary": "../data/maps/majority_calgary.jpg",
            "minority_calgary": "../data/maps/minority_calgary.jpg",
            "minority_other_cities": "../source_maps/minority_other_cities.jpg",
        },
    )
    polls = apportion_special_ballots(polls)

    print(f"\n{'─' * 70}")
    print("Skeleton complete. Implement the [STUB] functions to proceed.")
    print("Each stub has implementation guidance in its docstring.")
    print("\nNext step: build alberta_landmarks.csv from school district lists.")


if __name__ == "__main__":
    main()
